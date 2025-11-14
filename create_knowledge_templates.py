"""
创建企业知识库Excel模板
帮助用户了解每个文件应该包含什么内容
"""
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("需要安装 openpyxl: pip install openpyxl")
    exit(1)


def create_dept_capacity_template():
    """创建部门生产力配置模板"""
    
    wb = Workbook()
    
    # 样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # ========== 工作表1：软件部 ==========
    ws_sw = wb.active
    ws_sw.title = "软件部"
    
    headers_sw = ["资源类型", "资源名称", "配置规格", "数量", "能力评估", "备注"]
    ws_sw.append(headers_sw)
    
    sample_sw = [
        ["服务器", "编译服务器", "32核64G内存", "3台", "并行编译3个项目", ""],
        ["服务器", "测试服务器", "16核32G内存", "2台", "并行测试2个版本", ""],
        ["服务器", "代码仓库", "GitLab企业版", "1套", "无限用户", ""],
        ["开发工具", "Android Studio", "专业版授权", "无限制", "所有开发人员", ""],
        ["开发工具", "逻辑分析仪", "高速逻辑分析仪", "5台", "调试硬件信号", ""],
        ["开发工具", "示波器", "数字示波器", "3台", "调试模拟信号", ""],
        ["人员能力", "Android开发", "高级3人+中级2人", "5人", "3个模块/月", "包含UI+Framework+HAL"],
        ["人员能力", "驱动开发", "Linux内核工程师", "2人", "5个驱动/月", "网络+音频+显示等"],
        ["人员能力", "测试能力", "测试工程师", "2人", "全功能测试3天", "手动+自动化测试"]
    ]
    
    for row_data in sample_sw:
        ws_sw.append(row_data)
    
    for col_idx, header in enumerate(headers_sw, 1):
        col_letter = get_column_letter(col_idx)
        ws_sw.column_dimensions[col_letter].width = 20
        ws_sw[f"{col_letter}1"].font = header_font
        ws_sw[f"{col_letter}1"].fill = header_fill
        ws_sw[f"{col_letter}1"].alignment = header_alignment
    
    # ========== 工作表2：生产部 ==========
    ws_prod = wb.create_sheet("生产部")
    
    headers_prod = ["资源类型", "资源名称", "配置规格", "数量", "产能", "备注"]
    ws_prod.append(headers_prod)
    
    sample_prod = [
        ["生产线", "SMT贴片线", "高速贴片机", "2条", "500片/天/线", "理论产能1000片/天"],
        ["生产线", "组装线", "手工+半自动", "3条", "300台/天/线", "理论产能900台/天"],
        ["生产线", "测试工位", "自动化测试系统", "5个", "200台/天/工位", "理论产能1000台/天"],
        ["瓶颈分析", "当前瓶颈", "测试工位", "", "1000台/天", "测试工位限制"],
        ["瓶颈分析", "理论产能", "SMT限制", "", "1000台/天", "贴片线限制"],
        ["瓶颈分析", "实际产能", "考虑良率", "", "900台/天", "良率约90%"],
        ["质量控制", "IQC", "来料检验", "2人", "500批次/天", ""],
        ["质量控制", "IPQC", "过程检验", "3人", "巡检全线", ""],
        ["质量控制", "OQC", "出货检验", "2人", "1000台/天", ""]
    ]
    
    for row_data in sample_prod:
        ws_prod.append(row_data)
    
    for col_idx, header in enumerate(headers_prod, 1):
        col_letter = get_column_letter(col_idx)
        ws_prod.column_dimensions[col_letter].width = 20
        ws_prod[f"{col_letter}1"].font = header_font
        ws_prod[f"{col_letter}1"].fill = header_fill
        ws_prod[f"{col_letter}1"].alignment = header_alignment
    
    # 保存
    wb.save("config/组织/部门生产力配置.xlsx")
    print("[OK] 部门生产力配置.xlsx")


def create_supplier_template():
    """创建供应商信息模板"""
    
    wb = Workbook()
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # ========== 工作表1：芯片供应商 ==========
    ws_chip = wb.active
    ws_chip.title = "芯片供应商"
    
    headers_chip = ["供应商名称", "联系人", "邮箱", "芯片型号", "供货周期(天)", "MOQ(片)", "单价(元)", "技术支持", "备注"]
    ws_chip.append(headers_chip)
    
    sample_chip = [
        ["展锐(Unisoc)", "孟瑜", "yu.meng@unisoc.com", "T310", "30", "1000", "33", "提供", "主力芯片"],
        ["展锐(Unisoc)", "孟瑜", "yu.meng@unisoc.com", "T610", "30", "1000", "45", "提供", "高端芯片"],
        ["展锐(Unisoc)", "孟瑜", "yu.meng@unisoc.com", "T760", "45", "1000", "68", "提供", "5G芯片"],
        ["MTK(联发科)", "待配置", "", "MT6762", "25", "500", "28", "有限", "备选方案"]
    ]
    
    for row_data in sample_chip:
        ws_chip.append(row_data)
    
    for col_idx, header in enumerate(headers_chip, 1):
        col_letter = get_column_letter(col_idx)
        ws_chip.column_dimensions[col_letter].width = 15
        ws_chip[f"{col_letter}1"].font = header_font
        ws_chip[f"{col_letter}1"].fill = header_fill
        ws_chip[f"{col_letter}1"].alignment = header_alignment
    
    # ========== 工作表2：器件供应商 ==========
    ws_comp = wb.create_sheet("器件供应商")
    
    headers_comp = ["供应商名称", "联系人", "邮箱", "主要器件", "供货周期(天)", "库存水平", "价格稳定性", "备注"]
    ws_comp.append(headers_comp)
    
    sample_comp = [
        ["XXX电子", "待配置", "", "电阻电容", "3", "充足", "稳定", "常规器件"],
        ["XXX模块", "待配置", "", "WiFi模块", "15", "中等", "波动", "关键器件"],
        ["XXX屏幕", "待配置", "", "LCD显示屏", "20", "中等", "稳定", ""],
        ["XXX电池", "待配置", "", "锂电池", "10", "充足", "稳定", ""]
    ]
    
    for row_data in sample_comp:
        ws_comp.append(row_data)
    
    for col_idx, header in enumerate(headers_comp, 1):
        col_letter = get_column_letter(col_idx)
        ws_comp.column_dimensions[col_letter].width = 15
        ws_comp[f"{col_letter}1"].font = header_font
        ws_comp[f"{col_letter}1"].fill = header_fill
        ws_comp[f"{col_letter}1"].alignment = header_alignment
    
    wb.save("config/供应商/供应商信息.xlsx")
    print("[OK] 供应商信息.xlsx")


def create_workflow_template():
    """创建工作流程模板"""
    
    wb = Workbook()
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # ========== 工作表1：需求评审流程 ==========
    ws_req = wb.active
    ws_req.title = "需求评审流程"
    
    headers_req = ["步骤", "环节名称", "责任人", "审批人", "时限", "输入", "输出", "协作人员"]
    ws_req.append(headers_req)
    
    sample_req = [
        [1, "需求提出", "客户/市场部", "", "即时", "客户需求", "需求文档", "PM"],
        [2, "技术评审", "技术部经理", "徐智慧", "2工作日", "需求文档", "技术评审报告", "PM,HW工程师,SW工程师"],
        [3, "资源评审", "研发部经理", "陈琦", "1工作日", "技术评审报告", "资源评估报告", "各模块负责人"],
        [4, "立项决策", "总经理", "梁红生", "1工作日", "技术+资源评估", "立项批准", "技术部经理,研发部经理"],
        [5, "项目启动", "PM", "", "3工作日", "立项批准", "项目计划", "项目团队"]
    ]
    
    for row_data in sample_req:
        ws_req.append(row_data)
    
    for col_idx, header in enumerate(headers_req, 1):
        col_letter = get_column_letter(col_idx)
        ws_req.column_dimensions[col_letter].width = 18
        ws_req[f"{col_letter}1"].font = header_font
        ws_req[f"{col_letter}1"].fill = header_fill
        ws_req[f"{col_letter}1"].alignment = header_alignment
    
    # ========== 工作表2：开发流程 ==========
    ws_dev = wb.create_sheet("开发流程")
    
    headers_dev = ["步骤", "环节名称", "责任人", "审批人", "时限", "输入", "输出", "质量标准"]
    ws_dev.append(headers_dev)
    
    sample_dev = [
        [1, "需求分析", "PM", "技术负责人", "5工作日", "需求文档", "需求规格说明", "需求覆盖率100%"],
        [2, "设计评审", "技术负责人", "部门经理", "3工作日", "需求规格", "设计文档", "设计合理性"],
        [3, "编码实现", "开发人员", "Tech Lead", "按计划", "设计文档", "代码", "符合编码规范"],
        [4, "代码审查", "Tech Lead", "", "1工作日", "代码", "审查意见", "无严重问题"],
        [5, "单元测试", "开发人员", "", "与开发并行", "代码", "测试报告", "覆盖率>90%"],
        [6, "集成测试", "测试人员", "", "3工作日", "集成代码", "测试报告", "通过率>95%"],
        [7, "验收", "PM+客户", "", "5工作日", "测试通过", "验收报告", "客户满意"]
    ]
    
    for row_data in sample_dev:
        ws_dev.append(row_data)
    
    for col_idx, header in enumerate(headers_dev, 1):
        col_letter = get_column_letter(col_idx)
        ws_dev.column_dimensions[col_letter].width = 18
        ws_dev[f"{col_letter}1"].font = header_font
        ws_dev[f"{col_letter}1"].fill = header_fill
        ws_dev[f"{col_letter}1"].alignment = header_alignment
    
    wb.save("config/流程/研发工作流程.xlsx")
    print("[OK] 研发工作流程.xlsx")


def create_project_metadata_template():
    """创建项目元数据模板"""
    
    for project in ["G20", "889"]:
        metadata = {
            "project_code": project,
            "project_name": f"{project}智能终端" if project == "G20" else f"{project}项目",
            "customer": "九胜科技" if project == "G20" else "待配置",
            "pm": "张盛世" if project == "G20" else "待配置",
            "status": "量产" if project == "G20" else "开发中",
            "start_date": "2024-06-01",
            "target_date": "2025-02-01",
            "current_phase": "量产准备" if project == "G20" else "功能开发",
            "team": {
                "pm": "张盛世",
                "hw_lead": "付建章",
                "sw_lead": "邢鹏鹏",
                "test_lead": "测试组长",
                "members": ["李越征", "孙希鑫", "梁振东", "张海东"]
            },
            "milestones": [
                {"name": "需求确认", "date": "2024-06-15", "status": "已完成"},
                {"name": "方案设计", "date": "2024-07-01", "status": "已完成"},
                {"name": "硬件开发", "date": "2024-09-01", "status": "已完成"},
                {"name": "软件开发", "date": "2024-11-01", "status": "已完成"},
                {"name": "测试验证", "date": "2024-12-15", "status": "进行中"},
                {"name": "小批量试产", "date": "2025-01-15", "status": "计划中"},
                {"name": "正式量产", "date": "2025-02-01", "status": "计划中"}
            ],
            "resources": {
                "budget": "¥500,000",
                "personnel": "10人",
                "equipment": "标准开发设备"
            },
            "technical_stack": {
                "chip": "T310",
                "android_version": "Android 11",
                "kernel": "Linux 4.14",
                "wifi": "WiFi 5 (802.11ac)",
                "bluetooth": "BT 5.0",
                "cellular": "4G LTE Cat.4"
            },
            "risks": [
                {"risk": "芯片供货", "level": "低", "mitigation": "提前30天备货"},
                {"risk": "测试周期", "level": "中", "mitigation": "增加测试人手"}
            ],
            "dependencies": [
                {"item": "T310芯片", "supplier": "展锐", "lead_time": "30天"},
                {"item": "WiFi模块", "supplier": "XXX", "lead_time": "15天"}
            ]
        }
        
        import json
        with open(f"projects/{project}/metadata.json", 'w', encoding='utf-8') as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)
        
        print(f"[OK] projects/{project}/metadata.json")


def create_chip_knowledge_template():
    """创建芯片知识模板（Markdown）"""
    
    t310_template = """# T310 芯片技术文档

## 基本参数

- **芯片型号**: T310
- **供应商**: 展锐(Unisoc)
- **CPU**: Quad-core ARM Cortex-A75 @ 1.6GHz
- **GPU**: IMG PowerVR GE8300 @ 660MHz
- **内存**: 支持 LPDDR3/LPDDR4，最大4GB
- **存储**: 支持 eMMC 5.1，最大128GB

## 无线通信

### WiFi
- **标准**: 802.11 a/b/g/n/ac (WiFi 5)
- **频段**: 2.4GHz + 5GHz 双频
- **速率**: 最高433Mbps (11ac)
- **特性**: 
  - 支持2x2 MIMO
  - 支持WiFi Direct
  - 支持WiFi Hotspot

### 蓝牙
- **版本**: Bluetooth 5.0
- **BLE**: 支持
- **传输速率**: 2Mbps
- **共存**: 与WiFi共存（需注意干扰）

### 4G LTE
- **类别**: LTE Cat.4
- **下行**: 150Mbps
- **上行**: 50Mbps
- **频段**: B1/B3/B5/B7/B8/B20/B28/B38/B39/B40/B41

## 接口支持

- **显示**: MIPI DSI (最高1080p@60fps)
- **摄像头**: MIPI CSI (最高13MP)
- **音频**: I2S, PCM
- **USB**: USB 2.0 OTG
- **UART**: 3路
- **I2C**: 4路
- **SPI**: 3路
- **GPIO**: 50+

## 功耗特性

- **待机功耗**: < 5mA
- **通话功耗**: ~200mA
- **4G数据**: ~300mA
- **WiFi传输**: ~250mA

## 已知问题

1. **WiFi扫描延迟**
   - 问题: 首次WiFi扫描需要5-8秒
   - 影响: 用户体验
   - 解决: 固件v2.3已修复
   - 状态: 已解决

2. **BT与WiFi共存干扰**
   - 问题: BT和WiFi同时使用时可能互相干扰
   - 影响: 连接稳定性
   - 解决: 需要硬件隔离设计或使用共存算法
   - 状态: 需要设计注意

3. **GPS首次定位慢**
   - 问题: 冷启动定位需要30-60秒
   - 影响: 定位体验
   - 解决: 使用AGPS辅助定位
   - 状态: 需要软件配置

## 定制化选项

### GPIO配置
- 可配置的GPIO数量: 30+
- 功能复用: 支持

### 时钟配置
- 系统时钟: 可调整
- 外设时钟: 可配置

### 电源管理
- DCDC配置: 可调整
- LDO配置: 可调整
- 低功耗模式: 支持多种模式

## 开发资源

- **SDK**: Android 11 SDK
- **工具链**: GCC 8.3
- **调试工具**: JTAG, UART
- **文档**: 
  - Hardware Design Guide
  - Software Porting Guide
  - Driver Development Guide

## 技术支持

- **FAE支持**: 邮件/电话
- **响应时间**: 24小时内
- **现场支持**: 根据项目重要性
- **培训**: 可提供

## 参考案例

- G20项目: 使用T310，已量产
- 其他客户: 多个成功案例

## 注意事项

1. 硬件设计需遵循参考设计
2. WiFi天线设计影响性能
3. 电源设计需满足最大功耗
4. 散热设计需考虑最坏情况
"""
    
    with open("knowledge/suppliers/展锐(Unisoc)/chips/T310/technical_guide.md", 'w', encoding='utf-8') as f:
        f.write(t310_template)
    
    print("[OK] knowledge/suppliers/展锐(Unisoc)/chips/T310/technical_guide.md")


def main():
    print("=" * 70)
    print("创建企业知识库Excel模板")
    print("=" * 70)
    print()
    
    try:
        print("1. 创建部门生产力配置模板...")
        create_dept_capacity_template()
        
        print("\n2. 创建供应商信息模板...")
        create_supplier_template()
        
        print("\n3. 创建工作流程模板...")
        create_workflow_template()
        
        print("\n4. 创建项目元数据模板...")
        create_project_metadata_template()
        
        print("\n5. 创建芯片技术文档模板...")
        create_chip_knowledge_template()
        
        print()
        print("=" * 70)
        print("模板创建完成！")
        print("=" * 70)
        print()
        print("已创建的模板：")
        print("  1. config/组织/部门生产力配置.xlsx")
        print("  2. config/供应商/供应商信息.xlsx")
        print("  3. config/流程/研发工作流程.xlsx")
        print("  4. projects/G20/metadata.json")
        print("  5. projects/889/metadata.json")
        print("  6. knowledge/suppliers/展锐(Unisoc)/chips/T310/technical_guide.md")
        print()
        print("下一步：")
        print("  1. 打开这些模板文件")
        print("  2. 根据实际情况修改示例数据")
        print("  3. 添加您的技术文档到 knowledge/ 目录")
        print("  4. 运行 python data_inventory.py 查看完整度")
        print()
        
    except Exception as e:
        print(f"创建失败: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()

