# -*- coding: utf-8 -*-
"""
创建人员信息Excel模板 V5.3 - 支持组织架构和关系管理
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "人员信息"

# 设置表头样式
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

# V5.3 扩展表头：增加组织关系字段
headers = [
    # 基本信息
    "姓名", "邮箱", "人员类型", "公司", "职位",
    # 组织关系（V5.3新增）
    "所属部门", "直属领导", "负责部门", "负责客户",
    # 专业信息
    "技能", "工作经验", "学历", "技术水平",
    # 项目相关
    "负责项目", "沟通风格", "关注点", "职责", "备注"
]

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# 添加示例数据（展示组织架构）
examples = [
    # 高层领导
    ["梁红生", "lianghs@flyscale.cn", "top_leader", "飞思卡尔", "总经理",
     "", "", "技术部,市场部,运营部", "",  # 负责多个部门
     "", "20", "硕士", "",
     "", "", "", "公司战略和重大决策", "公司最高领导"],
    
    ["史国鹏", "shigp@flyscale.cn", "top_leader", "飞思卡尔", "副总经理",
     "", "", "研发部,测试部", "",
     "", "18", "硕士", "",
     "", "", "", "技术管理", "分管技术"],
    
    # 部门领导
    ["徐智慧", "xuzh@flyscale.cn", "department_leader", "飞思卡尔", "技术部经理",
     "技术部", "梁红生", "", "",  # 直属领导是梁红生
     "项目管理,需求分析", "10", "本科", "",
     "G20项目", "", "", "技术部整体管理", ""],
    
    ["陈琦", "chenqi@flyscale.cn", "department_leader", "飞思卡尔", "研发部经理",
     "研发部", "史国鹏", "", "",  # 直属领导是史国鹏
     "Android,嵌入式", "12", "硕士", "",
     "", "", "", "研发部管理", ""],
    
    # 项目经理（有负责客户）
    ["张盛世", "zhangsh@flyscale.cn", "pm", "飞思卡尔", "项目经理",
     "技术部", "徐智慧", "", "肖正伟",  # 负责客户肖正伟
     "项目管理", "8", "本科", "",
     "G20项目,889项目", "", "", "", "对接九胜客户"],
    
    ["罗宏", "luohong@flyscale.cn", "pm", "飞思卡尔", "项目经理",
     "技术部", "徐智慧", "", "Alex",  # 负责客户Alex
     "项目管理", "6", "本科", "",
     "", "", "", "", "对接台湾客户"],
    
    # 员工（有直属领导）
    ["邢鹏鹏", "xingpp@flyscale.cn", "employee", "飞思卡尔", "高级软件工程师",
     "研发部", "陈琦", "", "",  # 直属领导是陈琦
     "Android,C/C++,Linux驱动", "8", "本科", "",
     "G20项目,889项目", "", "", "Android应用和驱动开发", "技术能力强"],
    
    ["李越征", "liyz@flyscale.cn", "employee", "飞思卡尔", "软件工程师",
     "研发部", "陈琦", "", "",
     "Android,网络编程", "5", "本科", "",
     "G20项目", "", "", "网络模块开发", ""],
    
    ["孙希鑫", "sunxx@flyscale.cn", "employee", "飞思卡尔", "驱动工程师",
     "研发部", "陈琦", "", "",
     "Linux驱动,C", "6", "本科", "",
     "G20项目", "", "", "驱动开发", ""],
    
    # 客户
    ["肖正伟", "xiaozhengwei@szninetech.com", "customer", "深圳九胜科技", "CTO",
     "", "", "", "",
     "嵌入式,物联网,Android", "15", "博士", "expert",
     "G20项目,889项目", "技术导向", "质量,稳定性", "", "技术背景深厚"],
    
    ["Alex", "sales@maxcomm.com.tw", "customer", "台湾迈信", "销售经理",
     "", "", "", "",
     "", "", "", "basic",
     "", "商务导向", "交付时间,成本", "", ""],
    
    # 供应商
    ["孟瑜", "yu.meng@unisoc.com", "supplier", "展锐", "技术支持",
     "", "", "", "",
     "展锐芯片,T310,T610,BSP", "8", "硕士", "",
     "G20项目,889项目", "", "", "", "展锐主要联系人"],
]

for example in examples:
    ws.append(example)

# 调整列宽
column_widths = {
    'A': 12,  # 姓名
    'B': 30,  # 邮箱
    'C': 18,  # 人员类型
    'D': 15,  # 公司
    'E': 20,  # 职位
    'F': 15,  # 所属部门
    'G': 12,  # 直属领导
    'H': 25,  # 负责部门
    'I': 12,  # 负责客户
    'J': 35,  # 技能
    'K': 12,  # 工作经验
    'L': 10,  # 学历
    'M': 12,  # 技术水平
    'N': 25,  # 负责项目
    'O': 25,  # 沟通风格
    'P': 20,  # 关注点
    'Q': 25,  # 职责
    'R': 30,  # 备注
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 添加说明工作表
ws2 = wb.create_sheet("字段说明")
ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 55

instructions = [
    ["字段名", "说明"],
    ["姓名", "必填。人员姓名"],
    ["邮箱", "必填。邮箱地址，作为唯一标识"],
    ["人员类型", "必填。类型代码（见下方类型说明）"],
    ["公司", "可选。所属公司"],
    ["职位", "可选。职位/职务"],
    ["", ""],
    ["=== V5.3 组织关系字段 ===", ""],
    ["所属部门", "可选。员工/部门领导/PM所属的部门"],
    ["直属领导", "可选。员工的直接上级（填写姓名）"],
    ["负责部门", "可选。高层领导负责的部门（多个用逗号分隔）"],
    ["负责客户", "可选。PM负责对接的客户（填写姓名）"],
    ["", ""],
    ["技能", "可选。多个技能用逗号分隔，如：Python,Java,C++"],
    ["工作经验", "可选。年数，如：10"],
    ["学历", "可选。如：本科/硕士/博士"],
    ["技术水平", "可选（客户专用）。expert/intermediate/basic"],
    ["负责项目", "可选。多个项目用逗号分隔"],
    ["沟通风格", "可选（客户专用）。如：技术导向"],
    ["关注点", "可选（客户专用）。多个用逗号分隔"],
    ["职责", "可选。主要职责描述"],
    ["备注", "可选。其他备注信息"],
]

for row_idx, instruction in enumerate(instructions, 1):
    for col_idx, value in enumerate(instruction, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

# 添加类型说明工作表
ws3 = wb.create_sheet("人员类型说明")
ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 50

type_instructions = [
    ["人员类型", "类型代码", "说明"],
    ["高层领导", "top_leader", "公司高层，负责部门（填写'负责部门'字段）"],
    ["部门领导", "department_leader", "部门经理，有直属领导和所属部门"],
    ["项目经理", "pm", "项目经理，负责客户对接（填写'负责客户'字段）"],
    ["员工", "employee", "普通员工，有直属领导和所属部门"],
    ["客户", "customer", "客户联系人"],
    ["供应商", "supplier", "供应商联系人"],
]

for row_idx, instruction in enumerate(type_instructions, 1):
    for col_idx, value in enumerate(instruction, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

# 添加组织架构示例工作表
ws4 = wb.create_sheet("组织架构示例")
ws4.column_dimensions['A'].width = 60

org_example = [
    ["组织架构示例"],
    [""],
    ["公司组织架构："],
    [""],
    ["梁红生（总经理）- 高层领导"],
    ["├─ 负责部门：技术部、市场部、运营部"],
    ["│"],
    ["├─ 技术部"],
    ["│   ├─ 徐智慧（技术部经理）- 部门领导"],
    ["│   │   ├─ 张盛世（项目经理）- 负责客户：肖正伟"],
    ["│   │   └─ 罗宏（项目经理）- 负责客户：Alex"],
    ["│"],
    ["└─ 研发部（归史国鹏副总管理）"],
    ["    ├─ 陈琦（研发部经理）- 部门领导"],
    ["    │   ├─ 邢鹏鹏（员工）"],
    ["    │   ├─ 李越征（员工）"],
    ["    │   └─ 孙希鑫（员工）"],
    [""],
    ["关系说明："],
    ["1. 员工 → 直属领导 → 部门领导 → 高层领导"],
    ["2. 项目经理负责对接客户"],
    ["3. 通过姓名建立关联关系"],
]

for row_idx, line in enumerate(org_example, 1):
    cell = ws4.cell(row=row_idx, column=1, value=line[0])
    if row_idx == 1:
        cell.font = Font(bold=True, size=14)

# 保存文件
wb.save("persons/人员信息表_V5.3.xlsx")
print("[OK] Excel模板创建成功：persons/人员信息表_V5.3.xlsx")
print()
print("V5.3 新特性：")
print("1. 人员类型细分：top_leader/department_leader/pm/employee")
print("2. 组织关系字段：")
print("   - 所属部门：员工所在部门")
print("   - 直属领导：员工的直接上级")
print("   - 负责部门：高层领导负责的部门")
print("   - 负责客户：PM负责对接的客户")
print("3. 已包含组织架构示例")
print()
print("使用说明：")
print("1. 打开 persons/人员信息表_V5.3.xlsx")
print("2. 查看'组织架构示例'工作表了解结构")
print("3. 在'人员信息'工作表中填写数据")
print("4. 填写组织关系字段建立人员关联")
print("5. 保存并运行程序")

