"""
公司信息管理模块 - V6.0
管理和查询公司整体信息，构建"公司知识图谱"
"""
import os
import logging
from typing import Dict, Any, List, Optional
from pathlib import Path

# 可选依赖
try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("openpyxl 未安装，无法读取公司信息Excel。请运行: pip install openpyxl")

logger = logging.getLogger(__name__)


class CompanyManager:
    """公司信息管理器"""
    
    def __init__(self, config_root='config/公司'):
        """初始化公司管理器
        
        Args:
            config_root: 公司配置文件夹路径
        """
        self.config_root = Path(config_root)
        self.company_info = {}
        self.business_areas = {}
        self.products = {}
        self.departments = {}
        self.collaborations = []
        self.processes = {}
        
        self._load_company_info()
        self._load_department_info()
    
    def _load_company_info(self):
        """加载公司信息"""
        if not EXCEL_AVAILABLE:
            logger.error("无法加载公司信息：openpyxl未安装")
            return
        
        company_file = self.config_root / "公司信息.xlsx"
        if not company_file.exists():
            logger.info(f"公司信息文件不存在: {company_file}")
            return
        
        try:
            workbook = load_workbook(company_file, data_only=True)
            
            # 加载公司概况
            if "公司概况" in workbook.sheetnames:
                ws = workbook["公司概况"]
                headers = [cell.value for cell in ws[1]]
                
                for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                    for idx, value in enumerate(row):
                        if idx < len(headers) and headers[idx]:
                            self.company_info[headers[idx]] = value if value else ""
            
            # 加载业务领域
            if "业务领域" in workbook.sheetnames:
                ws = workbook["业务领域"]
                headers = [cell.value for cell in ws[1]]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # 业务领域不为空
                        area_name = row[0]
                        self.business_areas[area_name] = {
                            '描述': row[1] if len(row) > 1 else '',
                            '关键词': row[2].split(',') if len(row) > 2 and row[2] else [],
                            '负责部门': row[3] if len(row) > 3 else ''
                        }
            
            # 加载产品线
            if "产品线" in workbook.sheetnames:
                ws = workbook["产品线"]
                headers = [cell.value for cell in ws[1]]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[1]:  # 产品代码不为空
                        product_code = row[1]
                        self.products[product_code] = {
                            '系列': row[0] if row[0] else '',
                            '类型': row[2] if len(row) > 2 else '',
                            '主要客户': row[3] if len(row) > 3 else '',
                            '技术栈': row[4].split(',') if len(row) > 4 and row[4] else [],
                            '状态': row[5] if len(row) > 5 else ''
                        }
            
            logger.info(f"✅ 加载公司信息成功: {self.company_info.get('公司名称', '未知')}")
            logger.info(f"   业务领域: {len(self.business_areas)} 个")
            logger.info(f"   产品线: {len(self.products)} 个")
            
        except Exception as e:
            logger.error(f"加载公司信息失败: {str(e)}")
    
    def _load_department_info(self):
        """加载部门信息"""
        if not EXCEL_AVAILABLE:
            return
        
        dept_file = self.config_root / "部门信息.xlsx"
        if not dept_file.exists():
            logger.info(f"部门信息文件不存在: {dept_file}")
            return
        
        try:
            workbook = load_workbook(dept_file, data_only=True)
            
            # 加载部门职能
            if "部门职能" in workbook.sheetnames:
                ws = workbook["部门职能"]
                headers = [cell.value for cell in ws[1]]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # 部门名称不为空
                        dept_name = row[0]
                        self.departments[dept_name] = {
                            '负责人': row[1] if len(row) > 1 else '',
                            '职能描述': row[2] if len(row) > 2 else '',
                            '主要工作': row[3] if len(row) > 3 else '',
                            '协作部门': row[4].split(',') if len(row) > 4 and row[4] else [],
                            'KPI指标': row[5].split(',') if len(row) > 5 and row[5] else []
                        }
            
            # 加载部门协作关系
            if "部门协作关系" in workbook.sheetnames:
                ws = workbook["部门协作关系"]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1]:  # 发起部门和接收部门不为空
                        self.collaborations.append({
                            '发起部门': row[0],
                            '接收部门': row[1],
                            '协作类型': row[2] if len(row) > 2 else '',
                            '协作内容': row[3] if len(row) > 3 else '',
                            '频率': row[4] if len(row) > 4 else ''
                        })
            
            # 加载业务流程
            if "业务流程" in workbook.sheetnames:
                ws = workbook["业务流程"]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # 流程名称不为空
                        process_name = row[0]
                        self.processes[process_name] = {
                            '涉及部门': row[1].split(',') if len(row) > 1 and row[1] else [],
                            '关键步骤': row[2] if len(row) > 2 else '',
                            '责任人角色': row[3] if len(row) > 3 else '',
                            '时间要求': row[4] if len(row) > 4 else ''
                        }
            
            logger.info(f"✅ 加载部门信息成功")
            logger.info(f"   部门数: {len(self.departments)} 个")
            logger.info(f"   协作关系: {len(self.collaborations)} 条")
            logger.info(f"   业务流程: {len(self.processes)} 个")
            
        except Exception as e:
            logger.error(f"加载部门信息失败: {str(e)}")
    
    def get_company_overview(self):
        """获取公司概览信息"""
        return self.company_info
    
    def get_company_context_for_ai(self):
        """获取公司上下文信息（用于AI提示词）"""
        if not self.company_info:
            return ""
        
        context_parts = []
        context_parts.append("=" * 40)
        context_parts.append("【公司背景信息】")
        
        # 基本信息
        company_name = self.company_info.get('公司名称', '')
        if company_name:
            context_parts.append(f"公司：{company_name}")
        
        business = self.company_info.get('主营业务', '')
        if business:
            context_parts.append(f"主营业务：{business}")
        
        core_competence = self.company_info.get('核心竞争力', '')
        if core_competence:
            context_parts.append(f"核心竞争力：{core_competence}")
        
        # 业务领域
        if self.business_areas:
            context_parts.append("")
            context_parts.append("业务领域：")
            for area_name, area_info in self.business_areas.items():
                desc = area_info.get('描述', '')
                dept = area_info.get('负责部门', '')
                context_parts.append(f"  - {area_name}：{desc}（负责部门：{dept}）")
        
        # 产品线
        if self.products:
            context_parts.append("")
            context_parts.append("主要产品：")
            for product_code, product_info in list(self.products.items())[:5]:  # 最多5个
                product_type = product_info.get('类型', '')
                status = product_info.get('状态', '')
                context_parts.append(f"  - {product_code}：{product_type}（{status}）")
        
        context_parts.append("=" * 40)
        context_parts.append("")
        
        return '\n'.join(context_parts)
    
    def get_department_info(self, dept_name):
        """获取指定部门的详细信息"""
        return self.departments.get(dept_name)
    
    def get_department_context_for_ai(self, dept_name):
        """获取部门上下文信息（用于AI提示词）"""
        dept_info = self.get_department_info(dept_name)
        if not dept_info:
            return ""
        
        context_parts = []
        context_parts.append(f"【{dept_name}信息】")
        
        leader = dept_info.get('负责人', '')
        if leader:
            context_parts.append(f"负责人：{leader}")
        
        function = dept_info.get('职能描述', '')
        if function:
            context_parts.append(f"职能：{function}")
        
        work = dept_info.get('主要工作', '')
        if work:
            context_parts.append(f"主要工作：{work}")
        
        collab_depts = dept_info.get('协作部门', [])
        if collab_depts:
            context_parts.append(f"协作部门：{', '.join(collab_depts)}")
        
        kpis = dept_info.get('KPI指标', [])
        if kpis:
            context_parts.append(f"KPI指标：{', '.join(kpis)}")
        
        return '\n'.join(context_parts)
    
    def find_responsible_department(self, keyword):
        """根据关键词查找负责部门"""
        responsible_depts = []
        
        # 在业务领域中查找
        for area_name, area_info in self.business_areas.items():
            keywords = area_info.get('关键词', [])
            desc = area_info.get('描述', '')
            
            if keyword in keywords or keyword in desc:
                dept = area_info.get('负责部门', '')
                if dept and dept not in responsible_depts:
                    responsible_depts.append(dept)
        
        # 在部门主要工作中查找
        for dept_name, dept_info in self.departments.items():
            work = dept_info.get('主要工作', '')
            function = dept_info.get('职能描述', '')
            
            if keyword in work or keyword in function:
                if dept_name not in responsible_depts:
                    responsible_depts.append(dept_name)
        
        return responsible_depts
    
    def get_product_info(self, product_code):
        """获取产品信息"""
        return self.products.get(product_code)
    
    def has_company_info(self):
        """检查是否加载了公司信息"""
        return len(self.company_info) > 0 or len(self.departments) > 0


# 测试代码
if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO,
                       format='%(asctime)s - %(levelname)s - %(message)s')
    
    print("=" * 60)
    print("测试 CompanyManager")
    print("=" * 60)
    print()
    
    # 测试加载
    company_mgr = CompanyManager()
    
    if company_mgr.has_company_info():
        print("✓ 公司信息加载成功")
        print()
        
        # 测试获取公司概览
        print("【公司概览】")
        overview = company_mgr.get_company_overview()
        for key, value in overview.items():
            print(f"  {key}: {value}")
        print()
        
        # 测试获取AI上下文
        print("【AI上下文】")
        context = company_mgr.get_company_context_for_ai()
        print(context)
        
        # 测试部门查询
        print("【部门信息】")
        for dept_name in list(company_mgr.departments.keys())[:3]:
            print(f"\n{dept_name}:")
            dept_context = company_mgr.get_department_context_for_ai(dept_name)
            print(dept_context)
        
        # 测试责任部门查找
        print("\n【责任部门查找】")
        test_keywords = ["Android", "采购", "售后"]
        for keyword in test_keywords:
            depts = company_mgr.find_responsible_department(keyword)
            print(f"  关键词 '{keyword}' → 负责部门: {', '.join(depts)}")
    
    else:
        print("✗ 未找到公司信息")
        print()
        print("请先运行:")
        print("  python create_company_excel_template.py")

