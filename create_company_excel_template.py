"""
创建公司信息Excel模板（V6.0）
"""
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ openpyxl未安装，请运行: pip install openpyxl")
    exit(1)

def create_company_info_excel():
    """创建公司信息.xlsx"""
    
    wb = Workbook()
    
    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # ========== 工作表1：公司概况 ==========
    ws_overview = wb.active
    ws_overview.title = "公司概况"
    
    headers_overview = [
        "公司名称", "英文名称", "成立时间", "主营业务", "核心竞争力",
        "愿景", "核心价值观", "员工规模", "备注"
    ]
    
    ws_overview.append(headers_overview)
    
    # 示例数据
    sample_overview = [
        "飞思卡尔", "Flyscale", "2010-01-01", 
        "无线终端产品研发、生产与销售",
        "软件开发、硬件设计、快速交付能力",
        "成为行业领先的智能终端解决方案提供商",
        "创新、质量、客户至上、团队协作",
        "50-100人", "示例数据，请根据实际情况修改"
    ]
    
    ws_overview.append(sample_overview)
    
    # 设置样式
    for col_idx, header in enumerate(headers_overview, 1):
        col_letter = get_column_letter(col_idx)
        ws_overview.column_dimensions[col_letter].width = 20
        ws_overview[f"{col_letter}1"].font = header_font
        ws_overview[f"{col_letter}1"].fill = header_fill
        ws_overview[f"{col_letter}1"].alignment = header_alignment
    
    # ========== 工作表2：业务领域 ==========
    ws_business = wb.create_sheet("业务领域")
    
    headers_business = ["业务领域", "描述", "关键词", "负责部门"]
    ws_business.append(headers_business)
    
    sample_business = [
        ["软件开发", "Android系统、Linux驱动、应用开发", "Android,Linux,C/C++,Java", "研发部"],
        ["硬件设计", "原理图设计、PCB设计、硬件调试", "原理图,PCB,硬件", "研发部"],
        ["采购与生产", "芯片采购、器件采购、生产管理", "采购,生产,供应链", "采购部,生产部"],
        ["售后服务", "技术支持、维修服务、用户反馈", "售后,维修,客服", "售后服务部"]
    ]
    
    for row_data in sample_business:
        ws_business.append(row_data)
    
    for col_idx, header in enumerate(headers_business, 1):
        col_letter = get_column_letter(col_idx)
        ws_business.column_dimensions[col_letter].width = 30
        ws_business[f"{col_letter}1"].font = header_font
        ws_business[f"{col_letter}1"].fill = header_fill
        ws_business[f"{col_letter}1"].alignment = header_alignment
    
    # ========== 工作表3：产品线 ==========
    ws_products = wb.create_sheet("产品线")
    
    headers_products = ["产品系列", "产品代码", "产品类型", "主要客户", "技术栈", "状态"]
    ws_products.append(headers_products)
    
    sample_products = [
        ["智能终端", "G20", "4G智能终端", "九胜科技", "Android,T310,4G,WiFi,BT", "量产"],
        ["智能终端", "889", "4G智能终端", "某客户", "Android,T610,4G,WiFi,BT", "量产"],
        ["智能终端", "F7", "5G智能终端", "某客户", "Android,T7xx,5G,WiFi,BT", "研发中"],
        ["智能终端", "F3D", "4G智能终端", "某客户", "Android,T310,4G", "规划中"]
    ]
    
    for row_data in sample_products:
        ws_products.append(row_data)
    
    for col_idx, header in enumerate(headers_products, 1):
        col_letter = get_column_letter(col_idx)
        ws_products.column_dimensions[col_letter].width = 25
        ws_products[f"{col_letter}1"].font = header_font
        ws_products[f"{col_letter}1"].fill = header_fill
        ws_products[f"{col_letter}1"].alignment = header_alignment
    
    # 保存文件
    wb.save("config/公司/公司信息.xlsx")
    print("[OK] 公司信息.xlsx 创建成功")


def create_department_info_excel():
    """创建部门信息.xlsx"""
    
    wb = Workbook()
    
    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # ========== 工作表1：部门职能 ==========
    ws_dept = wb.active
    ws_dept.title = "部门职能"
    
    headers_dept = [
        "部门名称", "负责人", "职能描述", "主要工作", "协作部门", "KPI指标"
    ]
    
    ws_dept.append(headers_dept)
    
    sample_dept = [
        ["研发部", "陈琦", "软硬件研发、测试验证", 
         "Android开发、驱动开发、硬件设计、测试", 
         "技术部,采购部", 
         "项目按时交付率,Bug率,代码质量"],
        
        ["技术部", "徐智慧", "技术支持、项目管理", 
         "客户技术支持、项目管理、方案设计、售前支持", 
         "研发部,市场部", 
         "客户满意度,问题解决效率,项目成功率"],
        
        ["市场部", "梁红生", "市场推广、销售", 
         "市场调研、品牌推广、销售策略、渠道拓展", 
         "技术部,售后部", 
         "销售额,市场份额,品牌影响力"],
        
        ["采购部", "待配置", "采购管理、供应链", 
         "原材料采购、供应商管理、成本控制", 
         "研发部,生产部", 
         "采购成本,供应链稳定性,物料及时率"],
        
        ["生产部", "待配置", "生产制造、质量控制", 
         "产品生产、质量控制、库存管理", 
         "采购部,售后部", 
         "生产效率,产品合格率,交货及时率"],
        
        ["售后服务部", "待配置", "售后支持、维修", 
         "客户售后支持、维修、用户反馈收集", 
         "技术部,生产部", 
         "售后满意度,问题解决速度,返修率"]
    ]
    
    for row_data in sample_dept:
        ws_dept.append(row_data)
    
    for col_idx, header in enumerate(headers_dept, 1):
        col_letter = get_column_letter(col_idx)
        ws_dept.column_dimensions[col_letter].width = 25
        ws_dept[f"{col_letter}1"].font = header_font
        ws_dept[f"{col_letter}1"].fill = header_fill
        ws_dept[f"{col_letter}1"].alignment = header_alignment
    
    # ========== 工作表2：部门协作关系 ==========
    ws_collab = wb.create_sheet("部门协作关系")
    
    headers_collab = ["发起部门", "接收部门", "协作类型", "协作内容", "频率"]
    ws_collab.append(headers_collab)
    
    sample_collab = [
        ["技术部", "研发部", "需求传递", "客户需求、技术问题", "每日"],
        ["研发部", "技术部", "进度反馈", "项目进度、技术方案", "每周"],
        ["技术部", "市场部", "方案支持", "技术方案、售前支持", "按需"],
        ["市场部", "技术部", "客户信息", "客户需求、市场反馈", "每周"],
        ["采购部", "研发部", "器件确认", "芯片选型、器件采购", "按需"],
        ["研发部", "采购部", "需求提交", "BOM清单、器件需求", "按项目"],
        ["生产部", "研发部", "工艺反馈", "生产问题、设计优化", "按需"],
        ["研发部", "生产部", "工艺文档", "生产文档、测试标准", "按项目"]
    ]
    
    for row_data in sample_collab:
        ws_collab.append(row_data)
    
    for col_idx, header in enumerate(headers_collab, 1):
        col_letter = get_column_letter(col_idx)
        ws_collab.column_dimensions[col_letter].width = 20
        ws_collab[f"{col_letter}1"].font = header_font
        ws_collab[f"{col_letter}1"].fill = header_fill
        ws_collab[f"{col_letter}1"].alignment = header_alignment
    
    # ========== 工作表3：业务流程 ==========
    ws_process = wb.create_sheet("业务流程")
    
    headers_process = ["流程名称", "涉及部门", "关键步骤", "责任人角色", "时间要求"]
    ws_process.append(headers_process)
    
    sample_process = [
        ["产品研发流程", 
         "技术部,研发部,测试", 
         "需求分析→方案设计→开发实现→测试验证→发布上线",
         "项目经理", 
         "按项目计划"],
        
        ["客户服务流程", 
         "市场部,技术部,研发部,生产部,售后部", 
         "需求确认→方案设计→报价→研发→生产→交付→售后",
         "项目经理", 
         "按合同约定"],
        
        ["供应链流程", 
         "研发部,采购部,生产部", 
         "需求提出→采购询价→下单→入库→生产→出库",
         "采购经理", 
         "按生产计划"],
        
        ["售后服务流程", 
         "售后部,技术部,研发部", 
         "问题接收→问题分析→解决方案→实施→反馈→归档",
         "售后经理", 
         "24小时响应"]
    ]
    
    for row_data in sample_process:
        ws_process.append(row_data)
    
    for col_idx, header in enumerate(headers_process, 1):
        col_letter = get_column_letter(col_idx)
        ws_process.column_dimensions[col_letter].width = 30
        ws_process[f"{col_letter}1"].font = header_font
        ws_process[f"{col_letter}1"].fill = header_fill
        ws_process[f"{col_letter}1"].alignment = header_alignment
    
    # 保存文件
    wb.save("config/公司/部门信息.xlsx")
    print("[OK] 部门信息.xlsx 创建成功")


if __name__ == "__main__":
    print("=" * 60)
    print("创建公司信息Excel模板（V6.0）")
    print("=" * 60)
    print()
    
    try:
        create_company_info_excel()
        create_department_info_excel()
        
        print()
        print("=" * 60)
        print("✓ 所有模板创建成功")
        print("=" * 60)
        print()
        print("已创建文件：")
        print("  1. config/公司/公司信息.xlsx")
        print("  2. config/公司/部门信息.xlsx")
        print()
        print("下一步：")
        print("  1. 打开Excel文件")
        print("  2. 根据实际情况修改示例数据")
        print("  3. 保存文件")
        print("  4. 运行程序测试")
        print()
        
    except Exception as e:
        print(f"✗ 创建失败: {str(e)}")
        import traceback
        traceback.print_exc()

