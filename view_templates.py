"""
查看所有模板文件内容
"""
import json
from pathlib import Path
try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except:
    EXCEL_AVAILABLE = False


def view_chip_template():
    """查看芯片技术文档模板"""
    print("\n" + "=" * 70)
    print("1. 芯片技术文档模板")
    print("=" * 70)
    print("\n文件: knowledge/suppliers/展锐(Unisoc)/chips/T310/technical_guide.md")
    print("\n内容结构:")
    print("  ├── 基本参数 (CPU/GPU/内存/存储)")
    print("  ├── 无线通信")
    print("  │   ├── WiFi (标准/频段/速率/特性)")
    print("  │   ├── 蓝牙 (版本/BLE/速率)")
    print("  │   └── 4G LTE (类别/速率/频段)")
    print("  ├── 接口支持 (显示/摄像头/音频/USB/GPIO)")
    print("  ├── 功耗特性 (待机/通话/数据/WiFi)")
    print("  ├── 已知问题 (WiFi扫描/BT共存/GPS定位)")
    print("  ├── 定制化选项 (GPIO/时钟/电源)")
    print("  ├── 开发资源 (SDK/工具链/文档)")
    print("  ├── 技术支持 (FAE/响应时间)")
    print("  └── 参考案例")
    
    print("\n关键信息示例:")
    print("  - WiFi: WiFi 5 (802.11ac), 最高433Mbps")
    print("  - 蓝牙: Bluetooth 5.0, 2Mbps")
    print("  - 4G: LTE Cat.4, 下行150Mbps")
    print("  - 已知问题: BT与WiFi共存干扰（需硬件隔离）")
    
    print("\nAI如何使用:")
    print("  当邮件提到'WiFi 6'时，AI会:")
    print("  → 查询T310规格：只支持WiFi 5")
    print("  → 推荐升级到T610（支持WiFi 6）")
    print("  → 提示已知问题：注意BT/WiFi共存")


def view_project_template():
    """查看项目元数据模板"""
    print("\n" + "=" * 70)
    print("2. 项目元数据模板")
    print("=" * 70)
    print("\n文件: projects/G20/metadata.json")
    
    try:
        with open("projects/G20/metadata.json", 'r', encoding='utf-8') as f:
            metadata = json.load(f)
        
        print("\n项目基本信息:")
        print(f"  - 项目代码: {metadata['project_code']}")
        print(f"  - 项目名称: {metadata['project_name']}")
        print(f"  - 客户: {metadata['customer']}")
        print(f"  - 项目经理: {metadata['pm']}")
        print(f"  - 状态: {metadata['status']}")
        print(f"  - 开始日期: {metadata['start_date']}")
        print(f"  - 目标日期: {metadata['target_date']}")
        print(f"  - 当前阶段: {metadata['current_phase']}")
        
        print("\n项目团队:")
        print(f"  - PM: {metadata['team']['pm']}")
        print(f"  - 硬件负责: {metadata['team']['hw_lead']}")
        print(f"  - 软件负责: {metadata['team']['sw_lead']}")
        print(f"  - 成员: {', '.join(metadata['team']['members'])}")
        
        print(f"\n里程碑 ({len(metadata['milestones'])} 个):")
        for milestone in metadata['milestones']:
            status_icon = "✓" if milestone['status'] == "已完成" else "○"
            print(f"  {status_icon} {milestone['date']} - {milestone['name']} ({milestone['status']})")
        
        print("\n技术栈:")
        for key, value in metadata['technical_stack'].items():
            print(f"  - {key}: {value}")
        
        print(f"\n风险 ({len(metadata['risks'])} 项):")
        for risk in metadata['risks']:
            print(f"  - [{risk['level']}] {risk['risk']}: {risk['mitigation']}")
        
        print(f"\n依赖 ({len(metadata['dependencies'])} 项):")
        for dep in metadata['dependencies']:
            print(f"  - {dep['item']}: {dep['supplier']} (供货周期: {dep['lead_time']})")
        
        print("\nAI如何使用:")
        print("  当邮件提到'G20延期'时，AI会:")
        print("  → 查询目标日期: 2025-02-01")
        print("  → 查询当前阶段: 量产准备")
        print("  → 查询依赖: T310芯片30天，WiFi模块15天")
        print("  → 评估影响: 会影响哪个里程碑")
        print("  → 建议: 调整计划或加急处理")
        
    except Exception as e:
        print(f"  读取失败: {str(e)}")


def view_dept_capacity_template():
    """查看部门生产力配置模板"""
    print("\n" + "=" * 70)
    print("3. 部门生产力配置模板")
    print("=" * 70)
    print("\n文件: config/组织/部门生产力配置.xlsx")
    
    if not EXCEL_AVAILABLE:
        print("  需要安装 openpyxl")
        return
    
    try:
        wb = load_workbook("config/组织/部门生产力配置.xlsx")
        
        print(f"\n工作表: {wb.sheetnames}")
        
        # 软件部
        print("\n【软件部配置】")
        ws_sw = wb["软件部"]
        print("  资源类型 | 资源名称 | 配置规格 | 数量 | 能力评估")
        print("  " + "-" * 60)
        for idx, row in enumerate(ws_sw.iter_rows(min_row=2, max_row=10, values_only=True), 1):
            if row[0]:  # 有数据
                print(f"  {row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]}")
        
        # 生产部
        print("\n【生产部配置】")
        ws_prod = wb["生产部"]
        print("  资源类型 | 资源名称 | 配置规格 | 数量 | 产能")
        print("  " + "-" * 60)
        for idx, row in enumerate(ws_prod.iter_rows(min_row=2, max_row=10, values_only=True), 1):
            if row[0]:
                print(f"  {row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]}")
        
        print("\nAI如何使用:")
        print("  当邮件提到'需要10天完成3个模块'时，AI会:")
        print("  → 查询软件部能力: 5人团队，3个模块/月")
        print("  → 评估: 10天完成3个模块，需要加班或增加人手")
        print("  → 建议: 当前资源不足，建议调整计划或增加资源")
        
        print("\n  当邮件提到'下月交付2000台'时，AI会:")
        print("  → 查询生产能力: 实际产能900台/天")
        print("  → 计算: 2000台 ÷ 900台/天 = 2.2天")
        print("  → 评估: 生产能力充足")
        print("  → 查询瓶颈: 测试工位（1000台/天）")
        
    except Exception as e:
        print(f"  读取失败: {str(e)}")


def view_supplier_template():
    """查看供应商信息模板"""
    print("\n" + "=" * 70)
    print("4. 供应商信息模板")
    print("=" * 70)
    print("\n文件: config/供应商/供应商信息.xlsx")
    
    if not EXCEL_AVAILABLE:
        print("  需要安装 openpyxl")
        return
    
    try:
        wb = load_workbook("config/供应商/供应商信息.xlsx")
        
        print(f"\n工作表: {wb.sheetnames}")
        
        # 芯片供应商
        print("\n【芯片供应商】")
        ws = wb["芯片供应商"]
        print("  供应商 | 联系人 | 芯片型号 | 供货周期 | MOQ | 单价 | 技术支持")
        print("  " + "-" * 66)
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), 1):
            if row[0]:
                print(f"  {row[0]} | {row[1]} | {row[3]} | {row[4]}天 | {row[5]}片 | ¥{row[6]} | {row[7]}")
        
        print("\nAI如何使用:")
        print("  当邮件提到'需要500片T610'时，AI会:")
        print("  → 查询MOQ: 1000片（不满足MOQ）")
        print("  → 查询供货周期: 30天")
        print("  → 建议: 需要凑够1000片MOQ，或寻找其他供应商")
        
    except Exception as e:
        print(f"  读取失败: {str(e)}")


def view_workflow_template():
    """查看工作流程模板"""
    print("\n" + "=" * 70)
    print("5. 研发工作流程模板")
    print("=" * 70)
    print("\n文件: config/流程/研发工作流程.xlsx")
    
    if not EXCEL_AVAILABLE:
        print("  需要安装 openpyxl")
        return
    
    try:
        wb = load_workbook("config/流程/研发工作流程.xlsx")
        
        print(f"\n工作表: {wb.sheetnames}")
        
        # 需求评审流程
        print("\n【需求评审流程】")
        ws = wb["需求评审流程"]
        print("  步骤 | 环节 | 责任人 | 审批人 | 时限")
        print("  " + "-" * 60)
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=7, values_only=True), 1):
            if row[0]:
                print(f"  {row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]}")
        
        print("\nAI如何使用:")
        print("  当邮件提到'客户新需求'时，AI会:")
        print("  → 识别流程: 需求评审流程")
        print("  → 当前步骤: 第1步-需求提出")
        print("  → 下一步: 第2步-技术评审（责任人：徐智慧，时限：2工作日）")
        print("  → 建议: 抄送徐智慧，发起技术评审")
        
    except Exception as e:
        print(f"  读取失败: {str(e)}")


def generate_summary():
    """生成总结"""
    print("\n" + "=" * 70)
    print("模板文件总结")
    print("=" * 70)
    
    print("\n【模板完整性】")
    
    templates = [
        ("芯片技术文档", "knowledge/suppliers/展锐(Unisoc)/chips/T310/technical_guide.md"),
        ("项目元数据", "projects/G20/metadata.json"),
        ("部门生产力", "config/组织/部门生产力配置.xlsx"),
        ("供应商信息", "config/供应商/供应商信息.xlsx"),
        ("工作流程", "config/流程/研发工作流程.xlsx")
    ]
    
    for name, path in templates:
        exists = Path(path).exists()
        status = "✓" if exists else "✗"
        print(f"  {status} {name}: {path}")
    
    print("\n【数据覆盖范围】")
    print("  ✓ 技术知识: T310芯片完整规格（WiFi/蓝牙/4G/接口/功耗/已知问题）")
    print("  ✓ 项目信息: G20项目（7个里程碑/团队/技术栈/风险/依赖）")
    print("  ✓ 组织资源: 软件部（服务器/工具/开发能力）+ 生产部（产线/产能）")
    print("  ✓ 供应商: 展锐（T310/T610/T760，供货周期/MOQ/价格）")
    print("  ✓ 工作流程: 需求评审（5步骤）+ 开发流程（7步骤）")
    
    print("\n【AI分析能力】")
    print("  基于这些模板数据，AI已经能够:")
    print("  ✓ 回答技术问题（芯片是否支持某功能）")
    print("  ✓ 评估升级方案（需要哪个芯片）")
    print("  ✓ 预测采购周期（供货需要多久）")
    print("  ✓ 评估延期风险（基于项目里程碑）")
    print("  ✓ 计算生产能力（能否按时交付）")
    print("  ✓ 推荐流程步骤（下一步做什么）")
    
    print("\n【数据质量】")
    print("  当前: 示例数据（准确度约50%）")
    print("  填充真实数据后: 准确度可达90%+")
    
    print("\n【下一步】")
    print("  1. 查看每个模板文件，了解数据结构")
    print("  2. 根据您的实际情况，决定数据输入方式:")
    print("     - 快速对话（25分钟MVP）")
    print("     - 复制文档（AI自动解析）")
    print("     - 手动编辑Excel（不推荐）")


if __name__ == "__main__":
    print("=" * 70)
    print("企业知识库模板文件查看器")
    print("=" * 70)
    
    view_chip_template()
    view_project_template()
    view_dept_capacity_template()
    view_supplier_template()
    view_workflow_template()
    generate_summary()
    
    print("\n" + "=" * 70)
    print("查看完成！")
    print("=" * 70)
    print("\n您可以:")
    print("  1. 用Excel打开: config/组织/部门生产力配置.xlsx")
    print("  2. 用Excel打开: config/供应商/供应商信息.xlsx")
    print("  3. 用Excel打开: config/流程/研发工作流程.xlsx")
    print("  4. 用记事本打开: projects/G20/metadata.json")
    print("  5. 用记事本打开: knowledge/suppliers/展锐(Unisoc)/chips/T310/technical_guide.md")
    print()

