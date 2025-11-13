# -*- coding: utf-8 -*-
"""
创建人员信息Excel模板
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "人员信息"

# 设置表头样式
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

# 表头
headers = [
    "姓名", "邮箱", "类型", "公司", "职位",
    "技能", "工作经验", "学历", "技术水平",
    "负责项目", "沟通风格", "关注点", "职责", "备注"
]

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# 添加示例数据
examples = [
    ["肖正伟", "xiaozhengwei@szninetech.com", "customer", "深圳九胜科技", "技术总监",
     "嵌入式,物联网,Android", "15", "博士", "expert",
     "G20项目,889项目", "技术导向，喜欢深入讨论", "质量,稳定性", "", "技术背景深厚"],
    
    ["孟瑜", "yu.meng@unisoc.com", "supplier", "展锐", "技术支持",
     "展锐芯片,T310,T610,BSP", "8", "硕士", "",
     "G20项目,889项目", "", "", "", "展锐主要联系人"],
    
    ["梁红生", "lianghs@flyscale.cn", "leader", "飞思卡尔", "总经理",
     "", "20", "硕士", "",
     "", "", "", "公司战略和重大决策", "高层领导"],
    
    ["徐智慧", "xuzh@flyscale.cn", "pm", "飞思卡尔", "项目经理",
     "项目管理,需求分析", "10", "本科", "",
     "G20项目", "", "", "G20项目管理", "经验丰富"],
    
    ["邢鹏鹏", "xingpp@flyscale.cn", "employee", "飞思卡尔", "高级软件工程师",
     "Android,C/C++,Linux", "8", "本科", "",
     "G20项目,889项目", "", "", "Android应用和驱动", "技术能力强"],
]

for example in examples:
    ws.append(example)

# 调整列宽
column_widths = {
    'A': 12,  # 姓名
    'B': 30,  # 邮箱
    'C': 12,  # 类型
    'D': 15,  # 公司
    'E': 18,  # 职位
    'F': 35,  # 技能
    'G': 12,  # 工作经验
    'H': 10,  # 学历
    'I': 12,  # 技术水平
    'J': 25,  # 负责项目
    'K': 25,  # 沟通风格
    'L': 20,  # 关注点
    'M': 25,  # 职责
    'N': 30,  # 备注
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 添加说明工作表
ws2 = wb.create_sheet("填写说明")
ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 50

instructions = [
    ["字段", "说明"],
    ["姓名", "必填。人员姓名"],
    ["邮箱", "必填。邮箱地址，作为唯一标识"],
    ["类型", "必填。customer/supplier/leader/pm/employee"],
    ["公司", "可选。所属公司"],
    ["职位", "可选。职位/职务"],
    ["技能", "可选。多个技能用逗号分隔，如：Python,Java,C++"],
    ["工作经验", "可选。年数，如：10"],
    ["学历", "可选。如：本科/硕士/博士"],
    ["技术水平", "可选（客户专用）。expert/intermediate/basic"],
    ["负责项目", "可选。多个项目用逗号分隔"],
    ["沟通风格", "可选（客户专用）。如：技术导向"],
    ["关注点", "可选（客户专用）。多个用逗号分隔"],
    ["职责", "可选。主要职责描述"],
    ["备注", "可选。其他备注信息"],
]

for row_idx, instruction in enumerate(instructions, 1):
    for col_idx, value in enumerate(instruction, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

# 保存文件
wb.save("persons/人员信息表.xlsx")
print("[OK] Excel模板创建成功：persons/人员信息表.xlsx")
print()
print("使用说明：")
print("1. 打开 persons/人员信息表.xlsx")
print("2. 填写人员信息（参考示例数据）")
print("3. 保存文件")
print("4. 运行程序，人员信息会自动加载")

