# -*- coding: utf-8 -*-
"""
创建示例Excel文件
"""
from openpyxl import Workbook

# 创建工作簿
wb = Workbook()

# 第一个工作表：产品参数
ws1 = wb.active
ws1.title = "产品参数"
ws1.append(["参数名称", "规格", "实测值", "备注"])
ws1.append(["芯片型号", "展锐T310", "展锐T310", "OK"])
ws1.append(["CPU主频", "2.0GHz", "1.95GHz", "正常范围"])
ws1.append(["内存", "3GB LPDDR4X", "3GB", "OK"])
ws1.append(["存储", "32GB eMMC", "31.2GB", "系统占用"])
ws1.append(["屏幕", "4.0英寸 800x480", "4.0英寸", "OK"])
ws1.append(["电池", "3000mAh", "3000mAh", "OK"])
ws1.append(["摄像头", "500万像素", "500万像素", "OK"])

# 第二个工作表：测试数据
ws2 = wb.create_sheet("测试数据")
ws2.append(["测试项", "预期结果", "实际结果", "状态"])
ws2.append(["启动时间", "<30秒", "28秒", "通过"])
ws2.append(["通话质量", "清晰", "清晰", "通过"])
ws2.append(["网络切换", "无中断", "偶尔中断", "待优化"])
ws2.append(["GPS定位", "<30秒", "25秒", "通过"])
ws2.append(["电池续航", ">72小时", "68小时", "待优化"])

# 第三个工作表：问题追踪
ws3 = wb.create_sheet("问题追踪")
ws3.append(["问题ID", "问题描述", "优先级", "负责人", "状态"])
ws3.append(["BUG001", "网络切换时视频通话中断", "高", "李越征", "处理中"])
ws3.append(["BUG002", "长时间运行功耗偏高", "中", "孙希鑫", "处理中"])
ws3.append(["BUG003", "回声消除效果不理想", "中", "邢鹏鹏", "待处理"])

# 保存文件
wb.save("projects/G20/产品参数表.xlsx")
print("Excel文件创建成功：projects/G20/产品参数表.xlsx")

