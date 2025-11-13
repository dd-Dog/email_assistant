"""
人员信息文件加载器 - V5.2
支持从Excel表格和Markdown文件加载人员信息
"""
import os
import re
import logging
from typing import Dict, List, Optional

logger = logging.getLogger(__name__)

# 可选依赖
try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl 未安装，无法从Excel读取人员信息")


class PersonFileLoader:
    """人员信息文件加载器（支持Excel和Markdown）"""
    
    def __init__(self, persons_root='persons'):
        """初始化人员文件加载器
        
        Args:
            persons_root: 人员信息文件夹
        """
        self.persons_root = persons_root
        self.persons = {}
        self._load_all_persons()
    
    def _load_all_persons(self):
        """加载所有人员信息"""
        if not os.path.exists(self.persons_root):
            logger.info(f"人员信息目录不存在: {self.persons_root}")
            return
        
        # 优先从Excel加载（批量管理）
        excel_files = [f for f in os.listdir(self.persons_root) 
                      if f.endswith(('.xlsx', '.xls'))]
        
        for excel_file in excel_files:
            excel_path = os.path.join(self.persons_root, excel_file)
            persons_from_excel = self._load_from_excel(excel_path)
            self.persons.update(persons_from_excel)
        
        # 从Markdown/TXT文件加载（详细档案）
        for filename in os.listdir(self.persons_root):
            if filename.endswith(('.md', '.txt')):
                file_path = os.path.join(self.persons_root, filename)
                person = self._load_from_markdown(file_path)
                if person and 'email' in person:
                    self.persons[person['email']] = person
        
        if self.persons:
            logger.info(f"✅ 从文件加载了 {len(self.persons)} 个人员信息")
    
    def _load_from_excel(self, excel_path):
        """从Excel表格加载人员信息
        
        Excel格式：
        | 姓名 | 邮箱 | 类型 | 公司 | 职位 | 技能 | ... |
        
        Args:
            excel_path: Excel文件路径
            
        Returns:
            人员信息字典 {email: person_info}
        """
        if not EXCEL_AVAILABLE:
            logger.warning(f"跳过Excel {excel_path}（需要安装 openpyxl）")
            return {}
        
        try:
            workbook = load_workbook(excel_path, data_only=True)
            sheet = workbook.active
            
            persons = {}
            headers = []
            
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if row_idx == 1:
                    # 第一行是表头
                    headers = [str(cell).strip() if cell else '' for cell in row]
                    continue
                
                if not any(row):
                    continue
                
                # 构建人员信息
                person = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers) and headers[col_idx]:
                        header = headers[col_idx]
                        if value:
                            # 字段名映射
                            field_map = {
                                '姓名': 'name',
                                '邮箱': 'email',
                                '类型': 'type',
                                '公司': 'company',
                                '职位': 'role',
                                '技能': 'skills',
                                '工作经验': 'experience_years',
                                '学历': 'education',
                                '技术水平': 'technical_level',
                                '负责项目': 'projects',
                                '沟通风格': 'communication_style',
                                '关注点': 'priorities',
                                '职责': 'responsibilities',
                                '备注': 'notes'
                            }
                            
                            field = field_map.get(header, header)
                            value_str = str(value).strip()
                            
                            # 特殊处理：列表字段
                            if field in ['skills', 'projects', 'priorities']:
                                person[field] = [s.strip() for s in value_str.split(',') if s.strip()]
                            # 特殊处理：数字字段
                            elif field == 'experience_years':
                                try:
                                    person[field] = int(float(value_str))
                                except:
                                    person[field] = value_str
                            else:
                                person[field] = value_str
                
                # 必须有邮箱
                if 'email' in person:
                    persons[person['email']] = person
            
            logger.info(f"  从Excel加载了 {len(persons)} 个人员")
            return persons
            
        except Exception as e:
            logger.error(f"从Excel加载失败 {excel_path}: {str(e)}")
            return {}
    
    def _load_from_markdown(self, file_path):
        """从Markdown文件加载人员详细档案
        
        格式：
        # 张三
        
        - 邮箱: zhangsan@company.com
        - 类型: employee
        - 技能: Python, Java
        ...
        
        Args:
            file_path: Markdown文件路径
            
        Returns:
            人员信息字典
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
        except UnicodeDecodeError:
            try:
                with open(file_path, 'r', encoding='gbk') as f:
                    content = f.read()
            except Exception as e:
                logger.error(f"读取文件失败 {file_path}: {str(e)}")
                return None
        
        person = {}
        
        # 提取标题作为姓名
        title_match = re.search(r'^#\s+(.+)$', content, re.MULTILINE)
        if title_match:
            person['name'] = title_match.group(1).strip()
        
        # 提取键值对
        # 格式：- 键: 值 或 **键**: 值
        patterns = [
            r'^[-*]\s*\*?\*?(.+?)\*?\*?\s*[:：]\s*(.+)$',
            r'^\*\*(.+?)\*\*\s*[:：]\s*(.+)$'
        ]
        
        for pattern in patterns:
            for match in re.finditer(pattern, content, re.MULTILINE):
                key = match.group(1).strip()
                value = match.group(2).strip()
                
                # 字段名映射
                field_map = {
                    '姓名': 'name',
                    '邮箱': 'email',
                    '类型': 'type',
                    '公司': 'company',
                    '职位': 'role',
                    '技能': 'skills',
                    '工作经验': 'experience_years',
                    '学历': 'education',
                    '技术水平': 'technical_level',
                    '负责项目': 'projects',
                    '沟通风格': 'communication_style',
                    '关注点': 'priorities',
                    '职责': 'responsibilities',
                    '备注': 'notes'
                }
                
                field = field_map.get(key, key.lower())
                
                # 特殊处理：列表字段
                if field in ['skills', 'projects', 'priorities']:
                    person[field] = [s.strip() for s in value.split(',') if s.strip()]
                # 特殊处理：数字字段
                elif field == 'experience_years':
                    try:
                        person[field] = int(value.replace('年', ''))
                    except:
                        person[field] = value
                else:
                    person[field] = value
        
        return person if person else None
    
    def get_all_persons(self):
        """获取所有人员信息"""
        return self.persons
    
    def has_persons(self):
        """检查是否有人员信息"""
        return len(self.persons) > 0

