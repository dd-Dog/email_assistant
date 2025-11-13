# -*- coding: utf-8 -*-
"""
迁移工具：从config.json导出人员信息到Excel
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def migrate_config_to_excel():
    """从config.json导出人员信息到Excel"""
    
    print("=" * 60)
    print("配置迁移工具：config.json → Excel")
    print("=" * 60)
    print()
    
    # 读取config.json
    try:
        with open('config.json', 'r', encoding='utf-8') as f:
            config = json.load(f)
    except Exception as e:
        print(f"错误：无法读取config.json: {e}")
        return
    
    # 提取人员信息
    persons_data = []
    
    # 领导
    for email, name in config.get('leaders', {}).items():
        persons_data.append({
            'name': name,
            'email': email,
            'type': 'leader',
            'company': '飞思卡尔',  # 根据实际情况修改
        })
    
    # 项目经理
    for email, name in config.get('project_managers', {}).items():
        persons_data.append({
            'name': name,
            'email': email,
            'type': 'pm',
            'company': '飞思卡尔',
        })
    
    # 员工
    for email, name in config.get('employees', {}).items():
        persons_data.append({
            'name': name,
            'email': email,
            'type': 'employee',
            'company': '飞思卡尔',
        })
    
    # 客户
    for email, name in config.get('customers', {}).items():
        persons_data.append({
            'name': name,
            'email': email,
            'type': 'customer',
        })
    
    # 供应商
    for email, name in config.get('suppliers', {}).items():
        persons_data.append({
            'name': name,
            'email': email,
            'type': 'supplier',
        })
    
    if not persons_data:
        print("config.json中没有人员配置")
        return
    
    print(f"找到 {len(persons_data)} 个人员")
    print()
    
    # 创建Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "人员信息"
    
    # 设置表头样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 表头
    headers = [
        "姓名", "邮箱", "类型", "公司", "职位",
        "技能", "工作经验", "学历", "技术水平",
        "负责项目", "沟通风格", "关注点", "职责", "备注"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 填充数据
    for person in persons_data:
        row_data = [
            person.get('name', ''),
            person.get('email', ''),
            person.get('type', ''),
            person.get('company', ''),
            '',  # 职位
            '',  # 技能
            '',  # 工作经验
            '',  # 学历
            '',  # 技术水平
            '',  # 负责项目
            '',  # 沟通风格
            '',  # 关注点
            '',  # 职责
            ''   # 备注
        ]
        ws.append(row_data)
    
    # 调整列宽
    column_widths = {
        'A': 12, 'B': 30, 'C': 12, 'D': 15, 'E': 18,
        'F': 35, 'G': 12, 'H': 10, 'I': 12, 'J': 25,
        'K': 25, 'L': 20, 'M': 25, 'N': 30,
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 保存文件
    output_file = "persons/人员信息表_迁移.xlsx"
    wb.save(output_file)
    
    print(f"[OK] 成功导出到: {output_file}")
    print()
    print("下一步：")
    print("1. 打开 persons/人员信息表_迁移.xlsx")
    print("2. 补充详细信息（职位、技能等）")
    print("3. 重命名为 人员信息表.xlsx（覆盖模板）")
    print("4. 从config.json中删除人员配置（可选）")
    print()
    print("=" * 60)


if __name__ == '__main__':
    migrate_config_to_excel()

